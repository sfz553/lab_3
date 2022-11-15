import ast
from pprint import pprint
import os
import openpyxl
from openpyxl import load_workbook

def save():
    book = load_workbook('base.xlsx')
    sheet = book.active
    for i in range(len(massiv)):
        sheet.cell(row = i+1, column=1).value = massiv[i].name
        sheet.cell(row = i+1, column=2).value = massiv[i].colour
        sheet.cell(row = i+1, column=3).value = massiv[i].type
        sheet.cell(row = i+1, column=4).value = str(massiv[i].doors)
        sheet.cell(row = i+1, column=5).value = massiv[i].lightsst
    book.save('base.xlsx')
    book.close()
def load():
    saved = load_workbook('base.xlsx')
    sheet_saved = saved.active
    c = 1
    while sheet_saved.cell(row = c, column=1).value != None:
        #doorst = ast.literal_eval(sheet_saved.cell(row=c,column=4).value)
        #lightst = ast.literal_eval(sheet_saved.cell(row=c,column=5).value)
        massiv.append(Vehicle(name = sheet_saved.cell(row=c,column=1).value, colour=sheet_saved.cell(row=c,column=2).value, type=sheet_saved.cell(row=c,column=3).value, doors = ast.literal_eval(sheet_saved.cell(row=c,column=4).value), lightsst = sheet_saved.cell(row=c,column=5).value))
        c+=1
    saved.close()
def showlist():
    if len(massiv) == 0: print("Список пуст")
    else:
        for i in range(len(massiv)): print(f'{i+1}) {massiv[i].name}')
def showinfo():
    num = int(input('Для просмотра характеристик машины, выберите её номер из списка\n')) - 1
    print(f'Название: {massiv[num].name}\nЦвет: {massiv[num].colour}\nКатегория: {massiv[num].type}\nДвери: {massiv[num].doors}\nФары: {massiv[num].lightsst}\n')

def showall():
    for i in range(len(massiv)):
        print(f'Название: {massiv[i].name}\n Цвет: {massiv[i].colour}\n Категория: {massiv[i].type}\nДвери: {massiv[i].doors}\n Фары: {massiv[i].lightsst}\n')

class Doors:
    doors = ['Закрыто'] * 4
    def Open(self, num):
        self.doors[num - 1] = 'Открыто'
        print(f"Дверь {num} была открыта\n")
    def Close(self, num):
        self.doors[num - 1] = 'Закрыто'
        print(f"Дверь {num} была закрыта\n")
class Lights:
    def LightsUp(self):
        self.lightsst = 'Включены'
        print('Фары были включены')
    def LightsOff(self):
        self.lightsst = 'Выключены'
        print('Фары были выключены')
class Vehicle(Doors, Lights):

    def __init__(self, name=None, colour=None, type=None, doors = None, lightsst=None):
        self.name = name
        self.colour = colour
        self.type = type
        self.lightsst = lightsst
class Interface():

    def __init__(self):
        self.UserInterface()

    def UserInterface(self):
        while True:
            menu = input('1) Добавить машину\n2) Редактировать машину\n3) Список машин\n4) Сохранение\n5) Тестирование\n6) Полный список с характеристиками\n')

            # Добавить машину, menu = 1
            if menu == '1':
                massiv.append(Vehicle(name=input('Введите название машины\n'), colour=input('Введите цвет машины\n'), type=input('Введите категорию машины\n'), lightsst='Включены'))
                print('Каждая дверь закрыта по умолчанию\nФары включены по умолчанию')

            # Редактировать машину, menu = 2
            if menu == '2':
                showlist()
                i = int(input('Для редактирования машины, выберите её номер из списка\n')) - 1
                if i+1 > len(massiv):
                    print('Выберите номер существующей машины')
                else:

                    edit = int(input('Выберите характеристику для редактирования\n1) Название\n2) Цвет\n3) Категория\n4) Двери\n5) Фары\n6) Удалить машину\n'))

                    if edit == 1: massiv[i].name = input('Введите новое название:\n')
                    if edit == 2: massiv[i].colour = input('Введите новый цвет:\n')
                    if edit == 3: massiv[i].type = input('Введите новую категорию:\n')
                    if edit == 4:
                        print(massiv[i].doors)
                        dr = int(input("Введите номер двери, которую нужно открыть/закрыть\n"))
                        dow = int(input("1) Открыть дверь\n2) Закрыть дверь\n"))
                        if dow == 1:
                            massiv[i].Open(dr)
                        if dow == 2:
                            massiv[i].Close(dr)
                    if edit == 5:
                        uo = int(input('1) Включить фары\n2) Выключить фары\n'))
                        if uo == 1:
                            massiv[i].LightsUp()
                        if uo == 2:
                            massiv[i].LightsOff()
                    if edit == 6:
                        del massiv[i]
                        book = load_workbook('base.xlsx')
                        sheet = book.active
                        sheet.delete_rows(i + 1)
                        book.save('base.xlsx')
                        book.close()
                    else:
                        print('Выберите существующий пункт меню')


            # Список машин, menu = 3
            if menu == '3':
                showlist()
                showinfo()

            # Сохранение, menu = 4
            if menu == '4':
                save()

            # Тестирование, menu = 5
            if menu == '5':
                tm = input('Выберите, что нужно протестировать:\n1) Ручное управление файлом сохранения\n2) Вывод всех свойств и их названий в коде для всех объектов\n3) Тестирование удаления машины\n')

                if tm == '1':
                    os.system("start EXCEL.EXE base.xlsx")
                    break

                if tm == '2':
                    for i in range(len(massiv)):
                        pprint(vars(massiv[i]))

                if tm == '3':
                    showlist()
                    i = int(input('Введите номер машины для удаления\n')) - 1
                    book = load_workbook('base.xlsx')
                    sheet = book.active
                    print(sheet.cell(row = i+1, column=1).value, sheet.cell(row = i+1, column=2).value, sheet.cell(row = i+1, column=3).value, sheet.cell(row = i+1, column=4).value, sheet.cell(row = i+1, column=5).value)
                    x = input('Нажмите 1 для удаления\n')
                    if x == '1':
                        del massiv[i]
                        book = load_workbook('base.xlsx')
                        sheet = book.active
                        sheet.delete_rows(i + 1)
                        book.save('base.xlsx')
                        print(sheet.cell(row=i + 1, column=1).value, sheet.cell(row=i + 1, column=2).value,
                              sheet.cell(row=i + 1, column=3).value, sheet.cell(row=i + 1, column=4).value,
                              sheet.cell(row=i + 1, column=5).value)
                        book.close()

            if menu == '6':
                showall()



            else:
                print('Выберите существующий пункт меню')


massiv = []
load()
Interface()